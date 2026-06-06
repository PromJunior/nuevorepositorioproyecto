import csv
import logging
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

from fastapi.responses import StreamingResponse

from services.event_dispatcher import emit_report_generated as dispatch_report_generated

logger = logging.getLogger(__name__)


def _value(row: Any, key: str):
    if isinstance(row, dict):
        return row.get(key)
    return getattr(row, key, None)


def _serialize(value: Any):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return value


def _clean_filters(filters: dict[str, Any] | None) -> dict[str, Any]:
    return {
        key: _serialize(value)
        for key, value in (filters or {}).items()
        if value not in (None, "")
    }


def normalize_export_rows(rows: list[Any], columns: list[dict[str, str]]) -> list[dict[str, Any]]:
    return [
        {column["key"]: _serialize(_value(row, column["key"])) for column in columns}
        for row in rows
    ]


def build_csv_content(rows: list[Any], columns: list[dict[str, str]]) -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([column["label"] for column in columns])
    for row in normalize_export_rows(rows, columns):
        writer.writerow([row.get(column["key"], "") for column in columns])
    output.seek(0)
    return output.getvalue()


def build_csv_response(rows: list[Any], columns: list[dict[str, str]], filename: str) -> StreamingResponse:
    content = build_csv_content(rows, columns)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def build_excel_response(
    rows: list[Any],
    columns: list[dict[str, str]],
    filename: str,
    title: str,
    sheet_name: str,
    totals: dict[str, Any] | None = None,
) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Reporte"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)

    header_row = 3
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=column["label"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(12, min(36, len(column["label"]) + 8))

    for row in normalize_export_rows(rows, columns):
        ws.append([row.get(column["key"], "") for column in columns])

    if totals:
        ws.append([])
        ws.append(["TOTAL"] + [_serialize(totals.get(column["key"], "")) for column in columns[1:]])

    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def build_pdf_response(
    rows: list[Any],
    columns: list[dict[str, str]],
    filename: str,
    title: str,
    filters: dict[str, Any] | None = None,
    company: Any = None,
    totals: dict[str, Any] | None = None,
) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    muted = ParagraphStyle("muted", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    elements = []

    if company:
        company_name = getattr(company, "legal_name", None) or getattr(company, "trade_name", None)
        ruc = getattr(company, "ruc", None)
        if company_name:
            elements.append(Paragraph(company_name, styles["Heading2"]))
        if ruc:
            elements.append(Paragraph(f"RUC: {ruc}", muted))

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", muted))
    clean_filters = _clean_filters(filters)
    if clean_filters:
        elements.append(Paragraph("Filtros: " + " | ".join(f"{k}: {v}" for k, v in clean_filters.items()), muted))
    elements.append(Spacer(1, 0.35 * cm))

    table_rows = [[column["label"] for column in columns]]
    table_rows.extend([
        [str(row.get(column["key"], "")) for column in columns]
        for row in normalize_export_rows(rows, columns)
    ])
    if totals:
        table_rows.append(["TOTAL"] + [str(_serialize(totals.get(column["key"], ""))) for column in columns[1:]])

    table = Table(table_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def emit_export_event(
    module: str,
    format: str,
    filename: str,
    filters: dict[str, Any] | None,
    generated_by: str,
) -> dict[str, Any] | None:
    try:
        return dispatch_report_generated(
            module=module,
            format=format,
            filename=filename,
            filters=_clean_filters(filters),
            generated_by=generated_by,
        )
    except Exception:
        logger.exception("No se pudo emitir report.generated")
        return None
