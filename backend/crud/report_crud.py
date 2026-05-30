"""
Report CRUD — Consultas paginadas con filtros + generación de Excel/PDF.
Todas las queries son independientes para no mezclar lógica de dominio.
"""
from decimal import Decimal
from datetime import datetime, date
from io import BytesIO, StringIO
from typing import Optional
import csv

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import Date, func

from models.model import (
    Order, OrderItem, Payment, PaymentMethod,
    Purchase, PurchaseItem, PurchaseStatus,
    InventoryTransaction, InventoryTransactionType, Product,
    CashSession, User, AuditLog,
)
from crud.crm_crud import get_client_crm_rows


# ─── Helpers de fecha ─────────────────────────────────────────────────────────
def _dt_from(d: Optional[date]):
    return datetime(d.year, d.month, d.day, 0, 0, 0) if d else None

def _dt_to(d: Optional[date]):
    return datetime(d.year, d.month, d.day, 23, 59, 59) if d else None


# ─── REPORTE DE VENTAS ────────────────────────────────────────────────────────
def get_sales_report(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    user_id: Optional[int] = None,
    client_id: Optional[int] = None,
    payment_method_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 500,
) -> tuple[int, list]:
    q = db.query(Order).options(
        joinedload(Order.client),
        joinedload(Order.user),
        joinedload(Order.payment_order).joinedload(Payment.payment_method),
        joinedload(Order.order_items_order),
    )
    if date_from: q = q.filter(Order.order_date >= _dt_from(date_from))
    if date_to:   q = q.filter(Order.order_date <= _dt_to(date_to))
    if user_id:   q = q.filter(Order.user_id == user_id)
    if client_id: q = q.filter(Order.client_id == client_id)
    if payment_method_id:
        q = q.join(Payment, Payment.order_id == Order.id).filter(
            Payment.id_payment_method == payment_method_id
        )

    total = q.count()
    rows = q.order_by(Order.order_date.desc()).offset(skip).limit(limit).all()

    result = []
    for o in rows:
        pm = o.payment_order[0].payment_method.name_payment_method if o.payment_order else None
        result.append({
            "id": o.id,
            "order_date": o.order_date,
            "client_name": o.client.full_name if o.client else "Venta Mostrador",
            "seller_name": o.user.username if o.user else f"User#{o.user_id}",
            "items_count": len(o.order_items_order),
            "total_amount": Decimal(str(o.total_amount or 0)),
            "payment_method": pm,
        })
    return total, result


# ─── REPORTE DE COMPRAS ───────────────────────────────────────────────────────
def get_purchases_report(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    supplier_id: Optional[int] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 500,
) -> tuple[int, list]:
    q = db.query(Purchase).options(
        joinedload(Purchase.supplier),
        joinedload(Purchase.user),
        joinedload(Purchase.status),
        joinedload(Purchase.purchase_items),
    )
    if date_from:    q = q.filter(Purchase.purchase_date >= _dt_from(date_from))
    if date_to:      q = q.filter(Purchase.purchase_date <= _dt_to(date_to))
    if supplier_id:  q = q.filter(Purchase.supplier_id == supplier_id)
    if user_id:      q = q.filter(Purchase.user_id == user_id)
    if status:
        status_id = (
            db.query(PurchaseStatus.id)
            .filter(PurchaseStatus.name_status == status.upper())
            .scalar()
        )
        if status_id:
            q = q.filter(Purchase.status_id == status_id)

    total = q.count()
    rows = q.order_by(Purchase.purchase_date.desc()).offset(skip).limit(limit).all()

    return total, [
        {
            "id": p.id,
            "purchase_date": p.purchase_date,
            "supplier_name": p.supplier.company_name if p.supplier else f"Prov#{p.supplier_id}",
            "user_name": p.user.username if p.user else f"User#{p.user_id}",
            "invoice_number": p.invoice_number,
            "items_count": len(p.purchase_items),
            "total_amount": Decimal(str(p.total_amount or 0)),
            "status_name": p.status.name_status if p.status else None,
        }
        for p in rows
    ]


# ─── REPORTE KARDEX ───────────────────────────────────────────────────────────
def get_kardex_report(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    product_id: Optional[int] = None,
    transaction_type: Optional[str] = None,
    user_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 1000,
) -> tuple[int, list]:
    from crud.inventory_crud import get_transactions
    total, items = get_transactions(
        db=db,
        product_id=product_id,
        transaction_type=transaction_type,
        user_id=user_id,
        date_from=date_from,
        date_to=date_to,
        skip=skip,
        limit=limit,
    )
    return total, items


# ─── REPORTE DE CAJA ──────────────────────────────────────────────────────────
def _empty_daily_row(day: date) -> dict:
    return {
        "date": day,
        "stock_entries": 0,
        "stock_outputs": 0,
        "stock_adjustments": 0,
        "net_stock_movement": 0,
        "sales_count": 0,
        "sales_amount": Decimal("0.00"),
    }


def _as_date(value) -> date:
    return value.date() if hasattr(value, "date") else value


def _previous_balance(db: Session, tx: InventoryTransaction) -> Optional[int]:
    previous = (
        db.query(InventoryTransaction.balance_stock)
        .filter(
            InventoryTransaction.product_id == tx.product_id,
            InventoryTransaction.id < tx.id,
        )
        .order_by(InventoryTransaction.id.desc())
        .first()
    )
    return previous[0] if previous else None


def get_kardex_daily_summary(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    product_id: Optional[int] = None,
    user_id: Optional[int] = None,
    payment_method_id: Optional[int] = None,
    category_id: Optional[int] = None,
    source_type: Optional[str] = None,
    skip: int = 0,
    limit: int = 500,
) -> tuple[int, list]:
    rows_by_date: dict[date, dict] = {}

    tx_query = (
        db.query(InventoryTransaction)
        .join(
            InventoryTransactionType,
            InventoryTransaction.transaction_type_id == InventoryTransactionType.id,
        )
        .join(Product, Product.id == InventoryTransaction.product_id)
    )
    if date_from:
        tx_query = tx_query.filter(InventoryTransaction.created_at >= _dt_from(date_from))
    if date_to:
        tx_query = tx_query.filter(InventoryTransaction.created_at <= _dt_to(date_to))
    if product_id:
        tx_query = tx_query.filter(InventoryTransaction.product_id == product_id)
    if user_id:
        tx_query = tx_query.filter(InventoryTransaction.user_id == user_id)
    if category_id:
        tx_query = tx_query.filter(Product.category_id == category_id)
    if source_type:
        tx_query = tx_query.filter(InventoryTransaction.source_type == source_type)
    if payment_method_id:
        tx_query = (
            tx_query.join(
                Order,
                (InventoryTransaction.source_type == "orders")
                & (InventoryTransaction.source_id == Order.id),
            )
            .join(Payment, Payment.order_id == Order.id)
            .filter(Payment.id_payment_method == payment_method_id)
            .distinct()
        )

    transactions = tx_query.order_by(
        InventoryTransaction.created_at.asc(),
        InventoryTransaction.id.asc(),
    ).all()

    for tx in transactions:
        day = tx.created_at.date()
        row = rows_by_date.setdefault(day, _empty_daily_row(day))
        tx_type = (tx.transaction_type.name if tx.transaction_type else "").upper()
        qty = int(tx.quantity or tx.movement or 0)

        if tx_type == "ENTRADA":
            row["stock_entries"] += qty
        elif tx_type == "SALIDA":
            row["stock_outputs"] += qty
        elif tx_type == "AJUSTE":
            previous_balance = _previous_balance(db, tx)
            delta = int(tx.balance_stock or 0) - int(previous_balance or 0)
            row["stock_adjustments"] += delta
            if delta >= 0:
                row["stock_entries"] += delta
            else:
                row["stock_outputs"] += abs(delta)

    order_day = func.cast(Order.order_date, Date)
    if product_id or category_id:
        sales_query = (
            db.query(
                order_day.label("day"),
                func.count(func.distinct(Order.id)).label("sales_count"),
                func.coalesce(func.sum(OrderItem.sub_amount), 0).label("sales_amount"),
            )
            .join(OrderItem, OrderItem.order_id == Order.id)
            .join(Product, Product.id == OrderItem.product_id)
        )
    else:
        sales_query = db.query(
            order_day.label("day"),
            func.count(Order.id).label("sales_count"),
            func.coalesce(func.sum(Order.total_amount), 0).label("sales_amount"),
        )
    if date_from:
        sales_query = sales_query.filter(Order.order_date >= _dt_from(date_from))
    if date_to:
        sales_query = sales_query.filter(Order.order_date <= _dt_to(date_to))
    if user_id:
        sales_query = sales_query.filter(Order.user_id == user_id)
    if product_id:
        sales_query = sales_query.filter(OrderItem.product_id == product_id)
    if category_id:
        sales_query = sales_query.filter(Product.category_id == category_id)
    if payment_method_id:
        sales_query = sales_query.join(Payment, Payment.order_id == Order.id).filter(
            Payment.id_payment_method == payment_method_id
        )

    for day, sales_count, sales_amount in sales_query.group_by(order_day).all():
        day = _as_date(day)
        row = rows_by_date.setdefault(day, _empty_daily_row(day))
        row["sales_count"] = int(sales_count or 0)
        row["sales_amount"] = Decimal(str(sales_amount or 0))

    rows = []
    for row in rows_by_date.values():
        row["net_stock_movement"] = row["stock_entries"] - row["stock_outputs"]
        row["sales_amount"] = Decimal(str(row["sales_amount"] or 0))
        rows.append(row)

    rows.sort(key=lambda item: item["date"], reverse=True)
    total = len(rows)
    return total, rows[skip:skip + limit]


def get_cash_report(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    payment_method_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 200,
) -> tuple[int, list]:
    q = db.query(CashSession).options(joinedload(CashSession.user))
    if date_from: q = q.filter(CashSession.opening_time >= _dt_from(date_from))
    if date_to:   q = q.filter(CashSession.opening_time <= _dt_to(date_to))
    if user_id:   q = q.filter(CashSession.user_id == user_id)
    if status:    q = q.filter(CashSession.status == status.upper())
    if payment_method_id:
        q = (
            q.join(Order, Order.cash_session_id == CashSession.id)
            .join(Payment, Payment.order_id == Order.id)
            .filter(Payment.id_payment_method == payment_method_id)
            .distinct()
        )

    total = q.count()
    rows = q.order_by(CashSession.opening_time.desc()).offset(skip).limit(limit).all()

    return total, [
        {
            "id": s.id,
            "username": s.user.username if s.user else f"User#{s.user_id}",
            "opening_time": s.opening_time,
            "closing_time": s.closing_time,
            "opening_amount": Decimal(str(s.opening_amount or 0)),
            "expected_amount": Decimal(str(s.expected_amount)) if s.expected_amount else None,
            "closing_amount": Decimal(str(s.closing_amount)) if s.closing_amount else None,
            "difference": Decimal(str(s.difference)) if s.difference else None,
            "status": s.status,
        }
        for s in rows
    ]


# ─── AUDIT LOGS ──────────────────────────────────────────────────────────────
def get_audit_logs(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    user_id: Optional[int] = None,
    module: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
) -> tuple[int, list]:
    q = db.query(AuditLog).options(joinedload(AuditLog.user))
    if date_from: q = q.filter(AuditLog.created_at >= _dt_from(date_from))
    if date_to:   q = q.filter(AuditLog.created_at <= _dt_to(date_to))
    if user_id:   q = q.filter(AuditLog.user_id == user_id)
    if module:    q = q.filter(AuditLog.module == module)

    total = q.count()
    rows = q.order_by(AuditLog.created_at.desc()).offset(skip).limit(limit).all()

    return total, [
        {
            "id": r.id,
            "username": r.user.username if r.user else None,
            "module": r.module,
            "action": r.action,
            "entity": r.entity,
            "entity_id": r.entity_id,
            "description": r.description,
            "created_at": r.created_at,
        }
        for r in rows
    ]


def get_crm_report(
    db: Session,
    segment: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    user_id: Optional[int] = None,
    payment_method_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 500,
) -> tuple[int, list]:
    return get_client_crm_rows(
        db,
        segment=segment,
        date_from=date_from,
        date_to=date_to,
        user_id=user_id,
        payment_method_id=payment_method_id,
        skip=skip,
        limit=limit,
    )


def log_action(
    db: Session,
    user_id: Optional[int],
    module: str,
    action: str,
    entity: Optional[str] = None,
    entity_id: Optional[int] = None,
    description: Optional[str] = None,
):
    """Registra una acción en el log de auditoría. Llama db.commit() solo para el log."""
    try:
        entry = AuditLog(
            user_id=user_id,
            module=module,
            action=action,
            entity=entity,
            entity_id=entity_id,
            description=description,
        )
        db.add(entry)
        db.commit()
    except Exception:
        db.rollback()  # No propagar errores del log — nunca debe romper la operación principal


# ─── GENERADORES EXCEL ────────────────────────────────────────────────────────
def _style_header(ws, headers: list, col_widths: list):
    """Aplica estilo a la fila de cabecera."""
    from openpyxl.styles import Font, PatternFill, Alignment
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def generate_sales_excel(rows: list, title: str = "Reporte de Ventas") -> BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = ["ID", "Fecha", "Cliente", "Vendedor", "Ítems", "Total (S/.)", "Método Pago"]
    widths  = [8,    20,     30,       18,        8,      16,            20]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["id"],
            r["order_date"].strftime("%d/%m/%Y %H:%M") if r["order_date"] else "",
            r["client_name"],
            r["seller_name"],
            r["items_count"],
            float(r["total_amount"]),
            r["payment_method"] or "",
        ])
    # Fila de totales
    ws.append([])
    ws.append(["", "", "", "", "TOTAL", sum(float(r["total_amount"]) for r in rows), ""])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_purchases_excel(rows: list) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    headers = ["ID", "Fecha", "Proveedor", "Usuario", "N° Factura", "Ítems", "Total (S/.)", "Estado"]
    widths  = [8,    20,     30,          16,        16,             8,      16,             14]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["id"],
            r["purchase_date"].strftime("%d/%m/%Y %H:%M") if r["purchase_date"] else "",
            r["supplier_name"],
            r["user_name"],
            r["invoice_number"] or "",
            r["items_count"],
            float(r["total_amount"]),
            r["status_name"] or "",
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL", sum(float(r["total_amount"]) for r in rows), ""])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_kardex_excel(rows: list) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex"
    headers = ["ID", "Fecha", "Producto", "Categoría", "Tipo", "Concepto",
               "Cantidad", "Costo Unit.", "Saldo Stock", "Valor Saldo", "Usuario", "Origen"]
    widths  = [8,    20,     30,          18,           10,    25,
               9,          12,           12,             14,            16,       12]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["id"],
            r["created_at"].strftime("%d/%m/%Y %H:%M") if r.get("created_at") else "",
            r.get("product_name", ""),
            r.get("category_name", "") or "",
            r.get("transaction_type", ""),
            r.get("concept", ""),
            r.get("quantity", 0),
            float(r.get("unit_cost", 0)),
            r.get("balance_stock", 0),
            float(r.get("balance_value", 0)),
            r.get("username", ""),
            r.get("source_type", "") or "",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_kardex_daily_excel(rows: list, filters: Optional[dict] = None) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex Diario"
    headers = [
        "Fecha", "Entradas Stock", "Salidas Stock", "Ajustes",
        "Saldo Neto", "Ventas", "Monto Vendido (S/.)",
    ]
    widths = [14, 18, 18, 12, 14, 10, 20]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["date"].strftime("%d/%m/%Y") if r.get("date") else "",
            r.get("stock_entries", 0),
            r.get("stock_outputs", 0),
            r.get("stock_adjustments", 0),
            r.get("net_stock_movement", 0),
            r.get("sales_count", 0),
            float(r.get("sales_amount", 0)),
        ])
    ws.append([])
    ws.append([
        "TOTAL",
        sum(int(r.get("stock_entries", 0)) for r in rows),
        sum(int(r.get("stock_outputs", 0)) for r in rows),
        sum(int(r.get("stock_adjustments", 0)) for r in rows),
        sum(int(r.get("net_stock_movement", 0)) for r in rows),
        sum(int(r.get("sales_count", 0)) for r in rows),
        sum(float(r.get("sales_amount", 0)) for r in rows),
    ])

    if filters:
        filters_ws = wb.create_sheet("Filtros")
        filters_ws.append(["Filtro", "Valor"])
        for key, value in filters.items():
            if value not in (None, ""):
                filters_ws.append([key, str(value)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_cash_excel(rows: list) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caja"
    headers = ["ID", "Usuario", "Apertura", "Cierre", "Fondo Inicial",
               "Esperado", "Contado", "Diferencia", "Estado"]
    widths  = [8,    16,       20,       20,       14,
               14,       14,       12,          12]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["id"],
            r["username"],
            r["opening_time"].strftime("%d/%m/%Y %H:%M") if r["opening_time"] else "",
            r["closing_time"].strftime("%d/%m/%Y %H:%M") if r["closing_time"] else "",
            float(r["opening_amount"]),
            float(r["expected_amount"]) if r["expected_amount"] else "",
            float(r["closing_amount"]) if r["closing_amount"] else "",
            float(r["difference"]) if r["difference"] else "",
            r["status"],
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_crm_excel(rows: list) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM Clientes"
    headers = ["ID", "DNI", "Cliente", "Email", "Telefono", "Segmento", "Recency", "Frecuencia", "Monetary", "Ultima compra"]
    widths = [8, 14, 30, 30, 16, 14, 10, 12, 14, 18]
    _style_header(ws, headers, widths)
    for r in rows:
        ws.append([
            r["id"],
            r["dni"],
            r["full_name"],
            r["email"],
            r.get("phone") or "",
            r["segment"],
            r.get("recency_days") if r.get("recency_days") is not None else "",
            r["frequency"],
            float(r["monetary"]),
            r["last_purchase"].strftime("%d/%m/%Y") if r.get("last_purchase") else "",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_crm_csv(rows: list) -> StringIO:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "DNI", "Cliente", "Email", "Telefono", "Segmento", "Recency", "Frecuencia", "Monetary", "Ultima compra"])
    for r in rows:
        writer.writerow([
            r["id"],
            r["dni"],
            r["full_name"],
            r["email"],
            r.get("phone") or "",
            r["segment"],
            r.get("recency_days") if r.get("recency_days") is not None else "",
            r["frequency"],
            float(r["monetary"]),
            r["last_purchase"].strftime("%d/%m/%Y") if r.get("last_purchase") else "",
        ])
    output.seek(0)
    return output


# ─── GENERADORES PDF ──────────────────────────────────────────────────────────
def _build_pdf(
    title: str,
    headers: list,
    data_rows: list,
    landscape_mode: bool = True,
    filters: Optional[dict] = None,
) -> BytesIO:
    """Genera un PDF con tabla de datos usando reportlab."""
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    output = BytesIO()
    pagesize = rl_landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        output, pagesize=pagesize,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=8, textColor=colors.grey),
    ))
    if filters:
        filter_text = " | ".join(
            f"{key}: {value}" for key, value in filters.items() if value not in (None, "")
        )
        if filter_text:
            elements.append(Paragraph(
                f"Filtros: {filter_text}",
                ParagraphStyle("filters", parent=styles["Normal"], fontSize=8, textColor=colors.grey),
            ))
    elements.append(Spacer(1, 0.4*cm))

    # Tabla
    table_data = [headers] + data_rows
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),   colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),   colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),   "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),   9),
        ("ALIGN",       (0, 0), (-1, 0),   "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("FONTSIZE",    (0, 1), (-1, -1),  8),
        ("GRID",        (0, 0), (-1, -1),  0.3, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",  (0, 0), (-1, -1),  4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1),  5),
    ]))
    elements.append(t)

    doc.build(elements)
    output.seek(0)
    return output


def generate_sales_pdf(rows: list) -> BytesIO:
    headers = ["ID", "Fecha", "Cliente", "Vendedor", "Ítems", "Total S/.", "Método"]
    data = [
        [
            r["id"],
            r["order_date"].strftime("%d/%m/%Y %H:%M") if r["order_date"] else "",
            r["client_name"][:25],
            r["seller_name"],
            r["items_count"],
            f"S/. {float(r['total_amount']):.2f}",
            r["payment_method"] or "",
        ]
        for r in rows
    ]
    return _build_pdf("Reporte de Ventas", headers, data)


def generate_purchases_pdf(rows: list) -> BytesIO:
    headers = ["ID", "Fecha", "Proveedor", "Usuario", "Factura", "Ítems", "Total S/.", "Estado"]
    data = [
        [
            r["id"],
            r["purchase_date"].strftime("%d/%m/%Y") if r["purchase_date"] else "",
            r["supplier_name"][:20],
            r["user_name"],
            r["invoice_number"] or "",
            r["items_count"],
            f"S/. {float(r['total_amount']):.2f}",
            r["status_name"] or "",
        ]
        for r in rows
    ]
    return _build_pdf("Reporte de Compras", headers, data)


def generate_kardex_pdf(rows: list) -> BytesIO:
    headers = ["ID", "Fecha", "Producto", "Tipo", "Concepto", "Cant.", "Saldo", "Valor"]
    data = [
        [
            r["id"],
            r["created_at"].strftime("%d/%m/%Y") if r.get("created_at") else "",
            r.get("product_name", "")[:20],
            r.get("transaction_type", ""),
            r.get("concept", "")[:20],
            r.get("quantity", 0),
            r.get("balance_stock", 0),
            f"S/. {float(r.get('balance_value', 0)):.2f}",
        ]
        for r in rows
    ]
    return _build_pdf("Reporte Kardex", headers, data)


def generate_kardex_daily_pdf(rows: list, filters: Optional[dict] = None) -> BytesIO:
    headers = ["Fecha", "Entradas", "Salidas", "Ajustes", "Saldo Neto", "Ventas", "Monto"]
    data = [
        [
            r["date"].strftime("%d/%m/%Y") if r.get("date") else "",
            r.get("stock_entries", 0),
            r.get("stock_outputs", 0),
            r.get("stock_adjustments", 0),
            r.get("net_stock_movement", 0),
            r.get("sales_count", 0),
            f"S/. {float(r.get('sales_amount', 0)):.2f}",
        ]
        for r in rows
    ]
    totals = [
        "TOTAL",
        sum(int(r.get("stock_entries", 0)) for r in rows),
        sum(int(r.get("stock_outputs", 0)) for r in rows),
        sum(int(r.get("stock_adjustments", 0)) for r in rows),
        sum(int(r.get("net_stock_movement", 0)) for r in rows),
        sum(int(r.get("sales_count", 0)) for r in rows),
        f"S/. {sum(float(r.get('sales_amount', 0)) for r in rows):.2f}",
    ]
    return _build_pdf("Resumen Diario Kardex", headers, data + [totals], filters=filters)


def generate_cash_pdf(rows: list) -> BytesIO:
    headers = ["ID", "Usuario", "Apertura", "Cierre", "Fondo", "Esperado", "Contado", "Diferencia", "Estado"]
    data = [
        [
            r["id"],
            r["username"],
            r["opening_time"].strftime("%d/%m/%Y %H:%M") if r["opening_time"] else "",
            r["closing_time"].strftime("%d/%m/%Y %H:%M") if r["closing_time"] else "",
            f"S/. {float(r['opening_amount']):.2f}",
            f"S/. {float(r['expected_amount']):.2f}" if r["expected_amount"] else "",
            f"S/. {float(r['closing_amount']):.2f}" if r["closing_amount"] else "",
            f"S/. {float(r['difference']):.2f}" if r["difference"] else "",
            r["status"],
        ]
        for r in rows
    ]
    return _build_pdf("Reporte de Caja", headers, data)


def generate_crm_pdf(rows: list) -> BytesIO:
    headers = ["Cliente", "Segmento", "R", "F", "M", "Ultima compra"]
    data = [
        [
            r["full_name"][:28],
            r["segment"],
            r.get("recency_days") if r.get("recency_days") is not None else "",
            r["frequency"],
            f"S/. {float(r['monetary']):.2f}",
            r["last_purchase"].strftime("%d/%m/%Y") if r.get("last_purchase") else "",
        ]
        for r in rows
    ]
    return _build_pdf("Reporte CRM Clientes", headers, data)
